from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl as xl
import time
import random as r


#function to get games titles from IGN this will be used once!
def getGames():
    PATH = "C:\chromedriver.exe"
    driver = webdriver.Chrome(PATH)
    driver.get("https://www.ign.com/lists/top-100-games")
    games = driver.find_elements_by_class_name("list-item-heading")
    games_list = []
    for game in games :
        games_list.append(game.text)

    print("games:" , games_list)
    return games_list

#functions that set games titles from a Fn get games to excel sheets
def setGames(l):
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column
    #print(rows , "/" , columns)
    for i in range(2 , 101):
        print("Added game : " , l[i-1])
        cell = sheet.cell(i , 1)
        cell.value = l[i-1]

    wb.save("final1.xlsx")


#set Category rows
def setCategory():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        cell = sheet.cell(i , 3)
        cell.value = "Catgery" #call the fucntion here!

    wb.save("final1.xlsx")

#setting the price
def setPrice():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        cell = sheet.cell(i , 4)
        #print("Setting the Price for : " , title)
        price = "10"
        cell.value = price #call the fucntion here!

    wb.save("final1.xlsx")


#functions that setSpeces
def setSpeces ():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column

    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        print("Setting the Specs for : " , title)
        for j in range( 5 , 10):
            specs = sheet.cell(1 , j ).value
            cell = sheet.cell(i , j )
            #print(specs , " :")
            #value = input()
            cell.value = specs

    wb.save("final1.xlsx")


def setDes():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column

    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        print("Setting the Descriptions for : " , title)
        for j in range( 10 , 12):
            des = sheet.cell(1 , j ).value
            cell = sheet.cell(i , j )
            #print(des , " :")
            #value = input()
            cell.value = "it s a game !"

    wb.save("final1.xlsx")


def getImage(title):
    PATH = "C:\chromedriver.exe"
    driver = webdriver.Chrome(PATH)
    driver.get(f"https://www.google.com/search?q={title}")
    صور = driver.find_element_by_link_text("صور")
    صور.click()

    time.sleep(1)
    images =driver.find_elements_by_tag_name("img")
    i = r.randint(1,50)
    url = images[i].get_attribute("src")
    driver.close()
    return url


def setImages():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column

    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        print("Setting the images for : " , title)
        url = getImage(title)
        for j in range( 11 , 15):
            des = sheet.cell(1 , j ).value
            cell = sheet.cell(i , j )
            #print(des , " :")
            #value = input()
            cell.value = url
            print("setet images !")

    wb.save("final1.xlsx")

def setYtUrl():
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        print("Setting the YtUrl for : " , title)
        cell = sheet.cell(i,16)
        cell.value = 0


        wb.save("final1.xlsx")

def setConsole():
    conls = ["pc","ps3","ps4","xbox"]
    wb = xl.load_workbook("final1.xlsx")
    sheet = wb["Sheet1"]
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(2 , rows + 1 ):
        title = sheet.cell(i , 1).value
        cell = sheet.cell(i , 2)
        #print("Setting the Price for : " , title)
        price = conls[r.randint(0,3)]
        cell.value = price #call the fucntion here!

    wb.save("final1.xlsx")


#main fucntions
#games = getGames()
#setGames(games)
#setCategory()
#setPrice()
#setSpeces()
#setConsole()
#setDes()
#setImages()
#setYtUrl()