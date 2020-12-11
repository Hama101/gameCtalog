import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import random as r
import script as s

PATH = "C:\chromedriver.exe"
driver = webdriver.Chrome(PATH)
driver.get("http://127.0.0.1:8000/admin/")

username = driver.find_element_by_id("id_username")
username.send_keys("wajdi")
password = driver.find_element_by_id("id_password")
password.send_keys("wajdi2020")
password.send_keys(Keys.RETURN)

time.sleep(1)
add = driver.find_element_by_link_text('Gamess')
add.click()


wb = xl.load_workbook("final1.xlsx")
sheet = wb["Sheet1"]
rows = sheet.max_row
columns = sheet.max_column

for i in range(2, rows + 1 ):
    time.sleep(1)
    cell = sheet.cell(i,1)
    print("Adding to the data base this game : ",cell.value)
    addGames = driver.find_element_by_xpath('//*[@id="content-main"]/ul/li/a')
    addGames.click()
    time.sleep(1)
    inputs = driver.find_elements_by_tag_name("input")
    j=0
    while(True):
        j += 1
        if j == 2 :
            cell = sheet.cell(i,j+1)
            inputs[j].send_keys(cell.value)
            j += 1

        try:
            cell = sheet.cell(i,j+1)
            inputs[j].send_keys(cell.value)
            print(j)
        except :
            inputs[j].send_keys("error")
            print(j)

        if j == columns :
            break

    save = driver.find_element_by_xpath('//*[@id="games_form"]/div/div/input[1]')
    save.click()
    time.sleep(1)